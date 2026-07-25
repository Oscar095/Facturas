import io
import re

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from ..database import get_db
from ..models import Area, Factura, ReglaArea, Usuario
from ..schemas import AreaBase, AreaOut, ReglaAreaBase, ReglaAreaOut, ReglaAreaPatch
from ..security import requiere_permiso, usuario_actual
from ..services import reglas as reglas_svc

router = APIRouter(prefix="/api/areas", tags=["areas"])


def _limpiar_nit(valor: str | None) -> str | None:
    """Estandariza el NIT: solo dígitos (sin puntos/guiones); vacío -> None."""
    if not valor:
        return None
    s = re.sub(r"[.\s-]", "", str(valor).strip())
    return s or None


def _limpiar_texto(valor: str | None) -> str | None:
    v = (valor or "").strip()
    return v or None


@router.get("", response_model=list[AreaOut])
def listar_areas(db: Session = Depends(get_db), _: Usuario = Depends(usuario_actual)):
    return db.execute(select(Area).order_by(Area.nombre)).scalars().all()


@router.post("", response_model=AreaOut)
def crear_area(datos: AreaBase, db: Session = Depends(get_db),
               _: Usuario = Depends(requiere_permiso("administrar"))):
    if db.execute(select(Area).where(Area.nombre == datos.nombre)).scalar_one_or_none():
        raise HTTPException(400, "Ya existe un área con ese nombre")
    area = Area(**datos.model_dump())
    db.add(area)
    db.commit()
    db.refresh(area)
    return area


# ── reglas proveedor/ítem -> área ────────────────────────────────────────────────
@router.get("/reglas", response_model=list[ReglaAreaOut])
def listar_reglas(db: Session = Depends(get_db), _: Usuario = Depends(requiere_permiso("administrar"))):
    return db.execute(select(ReglaArea)).scalars().all()


@router.post("/reglas", response_model=ReglaAreaOut)
def crear_regla(datos: ReglaAreaBase, db: Session = Depends(get_db),
                _: Usuario = Depends(requiere_permiso("administrar"))):
    if db.get(Area, datos.area_id) is None:
        raise HTTPException(400, "El área indicada no existe")
    regla = ReglaArea(
        proveedor_nit=_limpiar_nit(datos.proveedor_nit),
        proveedor_nombre=_limpiar_texto(datos.proveedor_nombre),
        patron_item=_limpiar_texto(datos.patron_item),
        area_id=datos.area_id,
        responsable_id=datos.responsable_id,
    )
    db.add(regla)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "Ya existe una regla igual (mismo NIT, patrón y área)")
    db.refresh(regla)
    return regla


@router.patch("/reglas/{regla_id}", response_model=ReglaAreaOut)
def editar_regla(regla_id: int, datos: ReglaAreaPatch, db: Session = Depends(get_db),
                 _: Usuario = Depends(requiere_permiso("administrar"))):
    regla = db.get(ReglaArea, regla_id)
    if regla is None:
        raise HTTPException(404, "Regla no encontrada")
    cambios = datos.model_dump(exclude_unset=True)
    if "area_id" in cambios and db.get(Area, cambios["area_id"]) is None:
        raise HTTPException(400, "El área indicada no existe")
    if "proveedor_nit" in cambios:
        cambios["proveedor_nit"] = _limpiar_nit(cambios["proveedor_nit"])
    if "proveedor_nombre" in cambios:
        cambios["proveedor_nombre"] = _limpiar_texto(cambios["proveedor_nombre"])
    if "patron_item" in cambios:
        cambios["patron_item"] = _limpiar_texto(cambios["patron_item"])
    for campo, valor in cambios.items():
        setattr(regla, campo, valor)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(409, "Ya existe una regla igual (mismo NIT, patrón y área)")
    db.refresh(regla)
    return regla


@router.delete("/reglas/{regla_id}")
def eliminar_regla(regla_id: int, db: Session = Depends(get_db),
                   _: Usuario = Depends(requiere_permiso("administrar"))):
    regla = db.get(ReglaArea, regla_id)
    if regla is None:
        raise HTTPException(404, "Regla no encontrada")
    db.delete(regla)
    db.commit()
    return {"ok": True}


@router.post("/reglas/reaplicar")
def reaplicar_reglas(usar_ia: bool = False, db: Session = Depends(get_db),
                     _: Usuario = Depends(requiere_permiso("administrar"))):
    """Reaplica las reglas a las facturas SIN área (no toca las ya asignadas).

    `usar_ia=false` por defecto: solo asigna lo que las reglas deciden de forma
    determinística. Con `usar_ia=true`, para las que sigan ambiguas consulta a
    Claude (gasta créditos de API; queda auditado en eventos).
    """
    pendientes = db.execute(
        select(Factura).where(Factura.area_id.is_(None))
    ).scalars().all()
    asignadas = 0
    for f in pendientes:
        reglas_svc.asignar_area(db, f, usar_ia=usar_ia)
        if f.area_id is not None:
            reglas_svc.evaluar_completitud(db, f)
            asignadas += 1
    db.commit()
    return {"revisadas": len(pendientes), "asignadas": asignadas}


@router.post("/reglas/importar")
def importar_reglas(archivo: UploadFile = File(...), db: Session = Depends(get_db),
                    _: Usuario = Depends(requiere_permiso("administrar"))):
    """Importa el Excel proveedor/ítem -> área.

    Columnas esperadas (por nombre en la primera fila, sin distinción de mayúsculas):
      nit | proveedor_nit  (requerida)
      area                 (requerida, nombre del área; se crea si no existe)
      patron_item | item   (opcional)
      responsable_email    (opcional)
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(archivo.file.read()), read_only=True, data_only=True)
    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise HTTPException(400, "El archivo está vacío")

    encabezados = [str(c).strip().lower() if c is not None else "" for c in filas[0]]

    def col(*nombres):
        for n in nombres:
            if n in encabezados:
                return encabezados.index(n)
        return None

    i_nit = col("nit", "proveedor_nit")
    i_area = col("area", "área")
    i_patron = col("patron_item", "patron", "item", "ítem")
    i_resp = col("responsable_email", "responsable", "email")
    if i_nit is None or i_area is None:
        raise HTTPException(400, "El Excel debe tener columnas 'nit' y 'area'")

    areas_cache = {a.nombre.lower(): a for a in db.execute(select(Area)).scalars().all()}
    usuarios_cache = {u.email.lower(): u for u in db.execute(select(Usuario)).scalars().all()}

    creadas = 0
    for fila in filas[1:]:
        if not fila or fila[i_nit] is None:
            continue
        nit = str(fila[i_nit]).strip()
        nombre_area = str(fila[i_area]).strip()
        if not nit or not nombre_area:
            continue
        area = areas_cache.get(nombre_area.lower())
        if area is None:
            area = Area(nombre=nombre_area)
            db.add(area)
            db.flush()
            areas_cache[nombre_area.lower()] = area
        patron = str(fila[i_patron]).strip() if i_patron is not None and fila[i_patron] else None
        resp = None
        if i_resp is not None and fila[i_resp]:
            resp = usuarios_cache.get(str(fila[i_resp]).strip().lower())

        existente = db.execute(
            select(ReglaArea).where(
                ReglaArea.proveedor_nit == nit, ReglaArea.patron_item == patron
            )
        ).scalar_one_or_none()
        if existente:
            existente.area_id = area.id
            if resp:
                existente.responsable_id = resp.id
        else:
            db.add(ReglaArea(proveedor_nit=nit, patron_item=patron, area_id=area.id,
                             responsable_id=resp.id if resp else None))
            creadas += 1

    db.commit()
    return {"reglas_creadas": creadas}
