"""Analiza la hoja 'Detalle Proveedores': duplicados, proveedores con >1 área,
y cruce de nombre->NIT usando las hojas históricas (2024 / 2025-2026)."""
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

RUTA = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Relacion Facturas (2).xlsx")
wb = load_workbook(RUTA, read_only=True, data_only=True)


def normaliza(s):
    if not s:
        return ""
    s = str(s).strip().upper()
    s = re.sub(r"\s+", " ", s)
    return s


# ── Detalle Proveedores ──────────────────────────────────────────────────────
ws = wb["Detalle Proveedores"]
filas = list(ws.iter_rows(values_only=True))
proveedor_area = defaultdict(set)
proveedor_contacto = {}
total_filas = 0
for fila in filas[1:]:
    prov, area, contacto, cel = fila[0], fila[1], fila[2], fila[3]
    if not prov or not area:
        continue
    total_filas += 1
    p = normaliza(prov)
    proveedor_area[p].add(normaliza(area))
    proveedor_contacto[(p, normaliza(area))] = (contacto, cel)

print(f"Filas válidas: {total_filas}")
print(f"Proveedores únicos: {len(proveedor_area)}")
multi = {p: a for p, a in proveedor_area.items() if len(a) > 1}
print(f"Proveedores con MÁS DE 1 área: {len(multi)}")
for p, areas in list(multi.items())[:15]:
    print(f"  - {p}: {areas}")

print("\nÁreas distintas encontradas:")
todas_areas = set()
for a in proveedor_area.values():
    todas_areas |= a
print(sorted(todas_areas))

# ── Cruce nombre -> NIT usando hojas históricas ──────────────────────────────
nombre_a_nits = defaultdict(set)
for hoja in ("2024", "2025-2026"):
    ws2 = wb[hoja]
    encabezado = [normaliza(c) for c in next(ws2.iter_rows(values_only=True, max_row=1))[0]] \
        if False else None
    filas2 = ws2.iter_rows(values_only=True, max_row=20000)
    encabezados = [normaliza(c) if c else "" for c in next(filas2)]
    try:
        i_nit = encabezados.index("NIT")
        i_prov = encabezados.index("PROVEEDOR")
    except ValueError:
        print(f"Hoja {hoja}: no encontré columnas NIT/PROVEEDOR en {encabezados[:9]}")
        continue
    n = 0
    for fila in filas2:
        if fila[i_nit] is None or fila[i_prov] is None:
            continue
        nombre_a_nits[normaliza(fila[i_prov])].add(str(fila[i_nit]).strip())
        n += 1
    print(f"Hoja {hoja}: {n} filas con NIT+PROVEEDOR")

print(f"\nNombres únicos con NIT conocido (histórico): {len(nombre_a_nits)}")

# ¿Cuántos proveedores de 'Detalle Proveedores' cruzan directo por nombre exacto?
directos = 0
sin_match = []
for p in proveedor_area:
    if p in nombre_a_nits:
        directos += 1
    else:
        sin_match.append(p)
print(f"Cruce EXACTO por nombre: {directos} / {len(proveedor_area)}")
print(f"Sin cruce exacto: {len(sin_match)}")
print("Ejemplos sin cruce:", sin_match[:15])
