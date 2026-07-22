"""Resumen conciso de cada hoja: encabezados + 5 filas de muestra."""
import sys
from pathlib import Path

from openpyxl import load_workbook

RUTA = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Relacion Facturas (2).xlsx")

wb = load_workbook(RUTA, read_only=True, data_only=True)

for nombre in wb.sheetnames:
    ws = wb[nombre]
    print(f"\n=== {nombre} | filas={ws.max_row} cols={ws.max_column} ===")
    filas = list(ws.iter_rows(values_only=True, max_row=6))
    for i, fila in enumerate(filas):
        fila_corta = fila[:9]
        print(f"  {i}: {fila_corta}")
