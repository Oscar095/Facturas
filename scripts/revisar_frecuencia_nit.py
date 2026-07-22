import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

RUTA = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Relacion Facturas (2).xlsx")


def normaliza(s):
    return re.sub(r"\s+", " ", str(s).strip().upper()) if s else ""


def nit_valido(valor):
    s = re.sub(r"[.\s-]", "", str(valor).strip())
    return s if s.isdigit() and 6 <= len(s) <= 12 else None


wb = load_workbook(RUTA, read_only=True, data_only=True)
nombre_a_nits = defaultdict(Counter)
for hoja in ("2024", "2025-2026"):
    ws2 = wb[hoja]
    filas = ws2.iter_rows(values_only=True, max_row=20000)
    encabezados = [normaliza(c) if c else "" for c in next(filas)]
    i_nit, i_prov = encabezados.index("NIT"), encabezados.index("PROVEEDOR")
    for fila in filas:
        if fila[i_nit] is None or fila[i_prov] is None:
            continue
        nit = nit_valido(fila[i_nit])
        if nit:
            nombre_a_nits[normaliza(fila[i_prov])][nit] += 1

muestra = ["LOGISTICA INTEGRAL DE TRANSPORTE DE CARGA S.A.", "NEUMATICA Y CONTROLES S.A.S.",
           "DISCOVERY & PACK S.A.S", "BANDAS Y BANDAS SAS", "ROBINSON OCAMPO"]
for p in muestra:
    print(p, "->", nombre_a_nits.get(p))

# ¿cuántos casos se resuelven con "el NIT que aparece en más filas gana"
# con un margen claro (>=2x el segundo lugar)?
resueltos, no_resueltos = 0, 0
for p, contador in nombre_a_nits.items():
    if len(contador) <= 1:
        continue
    comunes = contador.most_common(2)
    top, segundo = comunes[0][1], (comunes[1][1] if len(comunes) > 1 else 0)
    if top >= 2 * max(segundo, 1):
        resueltos += 1
    else:
        no_resueltos += 1
print(f"\nAmbiguos totales: {resueltos + no_resueltos}")
print(f"Resueltos por mayoría clara (>=2x el segundo): {resueltos}")
print(f"Siguen empatados/dudosos: {no_resueltos}")
