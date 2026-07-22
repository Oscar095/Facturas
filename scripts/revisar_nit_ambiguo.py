import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

RUTA = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Relacion Facturas (2).xlsx")


def normaliza(s):
    return re.sub(r"\s+", " ", str(s).strip().upper()) if s else ""


def nit_valido(valor):
    s = re.sub(r"[.\s-]", "", str(valor).strip())
    return s if s.isdigit() and 6 <= len(s) <= 12 else None


wb = load_workbook(RUTA, read_only=True, data_only=True)

ws = wb["Detalle Proveedores"]
proveedores = set()
for fila in ws.iter_rows(values_only=True, min_row=2):
    if fila[0] and fila[1]:
        proveedores.add(normaliza(fila[0]))

nombre_a_nits = defaultdict(set)
for hoja in ("2024", "2025-2026"):
    ws2 = wb[hoja]
    filas = ws2.iter_rows(values_only=True, max_row=20000)
    encabezados = [normaliza(c) if c else "" for c in next(filas)]
    try:
        i_nit, i_prov = encabezados.index("NIT"), encabezados.index("PROVEEDOR")
    except ValueError:
        continue
    for fila in filas:
        if fila[i_nit] is None or fila[i_prov] is None:
            continue
        nit = nit_valido(fila[i_nit])
        if nit:
            nombre_a_nits[normaliza(fila[i_prov])].add(nit)

ambiguos = {p: nombre_a_nits[p] for p in proveedores if p in nombre_a_nits and len(nombre_a_nits[p]) > 1}
print(f"Total ambiguos: {len(ambiguos)}\n")
for p, nits in list(ambiguos.items())[:20]:
    print(f"{p}: {sorted(nits)}")
