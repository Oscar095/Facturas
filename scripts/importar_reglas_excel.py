"""Importa 'Detalle Proveedores' del Excel del usuario a reglas_area.

Cruza el nombre del proveedor contra las hojas históricas (2024, 2025-2026)
para obtener el NIT real (más confiable que el nombre en texto libre, ya que
es el mismo campo que entrega el portal Siesa). Reglas:
  - Proveedor con 1 sola área -> una fila (área asignada automáticamente).
  - Proveedor con 2+ áreas -> una fila por área, sin patrón (queda ambiguo
    hasta que la extracción de ítems o un admin decida). El operador puede
    luego cargar `patron_item` desde el panel para desambiguar.
  - Proveedor sin NIT cruzado -> se importa con NIT vacío, visible en el
    panel para completarlo a mano.

Uso: python scripts/importar_reglas_excel.py "Relacion Facturas (2).xlsx"
"""
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "backend"))
from openpyxl import load_workbook  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.database import SessionLocal  # noqa: E402
from app.models import Area, ReglaArea  # noqa: E402

RUTA = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Relacion Facturas (2).xlsx")


def normaliza(s):
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().upper())


wb = load_workbook(RUTA, read_only=True, data_only=True)

# ── 1. Proveedor -> {áreas} desde 'Detalle Proveedores' ────────────────────────
ws = wb["Detalle Proveedores"]
proveedor_areas = defaultdict(set)
for fila in ws.iter_rows(values_only=True, min_row=2):
    prov, area = fila[0], fila[1]
    if not prov or not area:
        continue
    proveedor_areas[normaliza(prov)].add(normaliza(area))

def nit_valido(valor) -> str | None:
    """Un NIT colombiano real es numérico (8-12 dígitos, sin el dígito de
    verificación). Descarta placeholders de texto como '-', 'N/A', etc."""
    s = re.sub(r"[.\s-]", "", str(valor).strip())
    return s if s.isdigit() and 6 <= len(s) <= 12 else None


# ── 2. Nombre -> NIT desde el histórico (por frecuencia, no solo presencia) ──
# El histórico tiene errores de tipeo/copiado (p.ej. NITs de terceros pegados
# por error en pocas filas); usamos el NIT que más se repite por nombre, y solo
# lo aceptamos si gana con margen claro (>=2x el segundo más frecuente).
nombre_a_nits = defaultdict(Counter)
for hoja in ("2024", "2025-2026"):
    ws2 = wb[hoja]
    filas = ws2.iter_rows(values_only=True, max_row=20000)
    encabezados = [normaliza(c) if c else "" for c in next(filas)]
    try:
        i_nit, i_prov = encabezados.index("NIT"), encabezados.index("PROVEEDOR")
    except ValueError:
        continue
    for fila in filas:
        if fila[i_nit] is None or fila[i_prov] is None:
            continue
        nit = nit_valido(fila[i_nit])
        if nit:
            nombre_a_nits[normaliza(fila[i_prov])][nit] += 1


def nit_por_mayoria(contador: Counter) -> str | None:
    if not contador:
        return None
    comunes = contador.most_common(2)
    top_nit, top_n = comunes[0]
    segundo_n = comunes[1][1] if len(comunes) > 1 else 0
    return top_nit if top_n >= 2 * max(segundo_n, 1) else None

# ── 3. Importar ──────────────────────────────────────────────────────────────
db = SessionLocal()
try:
    areas_cache = {a.nombre: a for a in db.execute(select(Area)).scalars().all()}

    def area_obj(nombre_area: str) -> Area:
        a = areas_cache.get(nombre_area)
        if a is None:
            a = Area(nombre=nombre_area)
            db.add(a)
            db.flush()
            areas_cache[nombre_area] = a
        return a

    # nombre "bonito" (tal como aparece en el Excel) para mostrar en el panel
    nombre_bonito = {}
    for fila in ws.iter_rows(values_only=True, min_row=2):
        if fila[0]:
            nombre_bonito.setdefault(normaliza(fila[0]), str(fila[0]).strip())

    creadas, con_nit, sin_nit, nit_ambiguo, ambiguos, omitidas_dup = 0, 0, 0, 0, 0, 0
    ya_insertado = set()  # (nit, area_id) -> evita duplicados si 2 nombres del Excel resuelven al mismo NIT
    for prov, areas in proveedor_areas.items():
        contador = nombre_a_nits.get(prov)
        if not contador:
            nit = None
            sin_nit += 1
        else:
            nit = nit_por_mayoria(contador)
            if nit:
                con_nit += 1
            else:
                nit_ambiguo += 1
        if len(areas) > 1:
            ambiguos += 1

        for nombre_area in areas:
            a = area_obj(nombre_area)
            if nit and (nit, a.id) in ya_insertado:
                omitidas_dup += 1
                continue
            if nit:
                ya_insertado.add((nit, a.id))
            db.add(ReglaArea(proveedor_nit=nit, proveedor_nombre=nombre_bonito.get(prov, prov),
                             patron_item=None, area_id=a.id, responsable_id=None))
            creadas += 1

    db.commit()
    print(f"Proveedores procesados: {len(proveedor_areas)}")
    print(f"  con NIT cruzado: {con_nit}")
    print(f"  SIN NIT en histórico (completar a mano): {sin_nit}")
    print(f"  NIT ambiguo en histórico, no se adivina (completar a mano): {nit_ambiguo}")
    print(f"  con más de un área candidata (quedan sin asignación automática): {ambiguos}")
    print(f"  omitidas por duplicado (mismo NIT+área ya insertado con otro nombre): {omitidas_dup}")
    print(f"Reglas creadas: {creadas}")
    print(f"Áreas: {sorted(areas_cache.keys())}")
finally:
    db.close()
