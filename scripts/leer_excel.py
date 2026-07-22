"""Inspecciona la estructura del Excel de relación de facturas -> responsable."""
import sys
from pathlib import Path

from openpyxl import load_workbook

RUTA = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("Relacion Facturas (2).xlsx")

wb = load_workbook(RUTA, read_only=True, data_only=True)
print("Hojas:", wb.sheetnames)

for nombre in wb.sheetnames:
    ws = wb[nombre]
    print(f"\n=== Hoja: {nombre} ({ws.max_row} filas x {ws.max_column} columnas) ===")
    filas = ws.iter_rows(values_only=True)
    for i, fila in enumerate(filas):
        print(i, fila)
        if i >= 12:
            print("...")
            break
